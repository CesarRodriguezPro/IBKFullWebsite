import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles import Color
from datetime import datetime

class vacationRequestBuilder:

    def __init__(self, employee, date_in, date_out, foreman):
        self.employee = employee
        self.date_in = date_in
        self.date_out = date_out
        self.foreman = foreman

    def run(self):

        wb = openpyxl.Workbook()
        ws = wb.active
        self.heading(ws)
        self.body(ws)
        self.document_sizes(ws)
        self.font_to_14(ws)
    
        wb.save('test_file.xlsx')
        wb.close()

    def heading(self, ws):
        ws.merge_cells("A1:F1")
        ws.merge_cells("A2:F2")
        ws.merge_cells("A3:F3")
        ws.merge_cells("A5:F5")

        ws['A1'] = 'IBK Construction Group'
        ws['A2'] = '617 Johnson Ave Brookly NY 11237'
        ws['A3'] = 'Tel: 718 418 0040, Fax: 718 732 2670'
        ws['A5'] = 'EMPLOYEE VACATION REQUEST FORM'

        # aligments 
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A5'].alignment = Alignment(horizontal='center')

        #colorig
        ws['A5'].fill = PatternFill(start_color='000000',end_color='000000',fill_type='solid')     
        ws['A5'].font = Font(color='ffffff')

    def body(self,ws):

        #spacing and formating
        ws.merge_cells("A9:B9")
        ws.merge_cells("C9:F9")
        ws.merge_cells("B15:C15")
        ws.merge_cells("E15:F15")
        ws.merge_cells("A19:F19")
        ws.merge_cells("A20:F20")
        ws.merge_cells("B23:D23")
        ws.merge_cells("B25:D25")

        # static content
        square = '\u25A2'
        ws['B7']  = 'Date:'
        ws['A9']  = 'Employee:'
        ws['A13'] = 'Vacation Date Requested:'
        ws['A15'] = 'From'
        ws['D15'] = 'To:'
        ws['A17'] = square +'Has been approved and will be paid on the check for which the vacation days occur'
        ws['A18'] = square + 'has not been approved'
        ws['A23'] = 'By:'
        ws['E23'] = 'Supervisor'
        ws['A25'] = 'By'
        ws['B29'] = 'Controller'

        #underlines and inputs
        ws['C7']  = datetime.strftime(datetime.today(), '%m/%d/%Y')
        ws['C9']  = self.employee
        ws['B15'] = self.date_in
        ws['E15'] = self.date_out 
        ws['B23'] = self.foreman
        ws['C7'].border   = self.underline(ws)
        ws['C9'].border   = self.underline(ws)
        ws['B15'].border  = self.underline(ws)
        ws['E15'].border  = self.underline(ws) 
        ws['A19'].border  = self.underline(ws)
        ws['A20'].border  = self.underline(ws)
        ws['B23'].border  = self.underline(ws)
        ws['B25'].border  = self.underline(ws)
        ws['A29'].border  = self.underline(ws)

        # font formating


    def document_sizes(self, ws): 
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 7
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

    def underline(self, ws):
        border_style = Border(bottom=Side(border_style='thin'))
        return border_style
    
    def font_to_14(self, ws):

        list_rows = [1,2,3,5,7,9,15,23,7,9,15,19,20,23,25,29]
        for n in list_rows:
            ws[f'A{n}'].font = Font(size=14)
            ws[f'B{n}'].font = Font(size=14)
            ws[f'C{n}'].font = Font(size=14)
            ws[f'D{n}'].font = Font(size=14)
            ws[f'E{n}'].font = Font(size=14)
            ws[f'F{n}'].font = Font(size=14)

def convert_to_pdf(location):
    file_dir = os.path.abspath(os.path.join(DIR_PATH, location))
    list_file_names = os.listdir(file_dir)
    for x in list_file_names:
        fullname = os.path.abspath(os.path.join(file_dir, x))
        print(fullname)
        os.system(f'"C:\Program Files\LibreOffice\program\soffice.bin"  --convert-to pdf {fullname} --outdir {file_dir}')



if __name__ == "__main__":
    active = vacationRequestBuilder(employee='Cesar rodriguez', date_in='2019-04-12', date_out='2019-05-12', foreman='Tyrone Mclance')
    active.run()